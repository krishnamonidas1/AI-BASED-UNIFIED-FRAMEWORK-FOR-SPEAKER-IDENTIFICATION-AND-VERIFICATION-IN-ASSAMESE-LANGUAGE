"""
export_splits.py
Generates dataset_splits.xlsx showing exactly which files went into
training and testing, organised by speaker.

Run: python export_splits.py
Output: results/dataset_splits.xlsx
"""

import os, glob, re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import LabelEncoder
from sklearn.model_selection import train_test_split

import config

# ── Reproduce exact same split as data_loader.py ─────────────────────────────

def _speaker_label(folder_name):
    num = re.search(r"\d+", folder_name)
    return f"Speaker_{int(num.group()):02d}" if num else folder_name


def collect_records():
    records = []
    for spk_dir in sorted(os.listdir(config.DATASET_ROOT)):
        full = os.path.join(config.DATASET_ROOT, spk_dir)
        if not os.path.isdir(full):
            continue
        label = _speaker_label(spk_dir)
        for wav in sorted(glob.glob(os.path.join(full, "*.wav"))):
            records.append((os.path.basename(wav), label, spk_dir))
    return records


def get_splits():
    records = collect_records()
    filenames = [r[0] for r in records]
    labels    = [r[1] for r in records]
    folders   = [r[2] for r in records]

    le = LabelEncoder()
    y  = le.fit_transform(labels)

    idx = np.arange(len(records))
    idx_tr, idx_te = train_test_split(
        idx, test_size=config.TEST_SIZE,
        random_state=config.RANDOM_STATE, stratify=y)

    train_set = set(idx_tr)
    rows = []
    for i, (fname, label, folder) in enumerate(records):
        split = "Train" if i in train_set else "Test"
        rows.append({
            "Speaker Label": label,
            "Original Folder": folder,
            "File Name": fname,
            "Split": split,
        })
    # Sort: speaker → split → filename
    rows.sort(key=lambda r: (r["Speaker Label"], r["Split"], r["File Name"]))
    return rows, le


# ── Styles ────────────────────────────────────────────────────────────────────

def _side():
    return Side(style="thin", color="D0D0D0")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
TRAIN_FILL   = PatternFill("solid", fgColor="E8F5E9")   # light green
TEST_FILL    = PatternFill("solid", fgColor="E3F2FD")   # light blue
SPK_FILL     = PatternFill("solid", fgColor="F5F5F5")   # light grey stripe
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="Arial", bold=True, size=14, color="1F3864")
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")


def _write_cell(ws, row, col, value, font=None, fill=None,
                alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:         cell.font        = font
    if fill:         cell.fill        = fill
    if alignment:    cell.alignment   = alignment
    if border:       cell.border      = border
    if number_format: cell.number_format = number_format
    return cell


# ── Sheet 1 – Full Split Detail ───────────────────────────────────────────────

def build_detail_sheet(wb, rows, le):
    ws = wb.active
    ws.title = "Full Split Detail"
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:E1")
    _write_cell(ws, 1, 1,
                "Assamese Speaker Dataset — Train / Test Split",
                font=TITLE_FONT, alignment=CENTER)
    ws.row_dimensions[1].height = 32

    # Sub-info row
    ws.merge_cells("A2:E2")
    info = (f"Total files: {len(rows)}  |  "
            f"Train: {sum(1 for r in rows if r['Split']=='Train')}  |  "
            f"Test: {sum(1 for r in rows if r['Split']=='Test')}  |  "
            f"Split ratio: {int((1-config.TEST_SIZE)*100)}/{int(config.TEST_SIZE*100)}  |  "
            f"Speakers: {len(le.classes_)}  |  "
            f"Random state: {config.RANDOM_STATE}")
    _write_cell(ws, 2, 1, info,
                font=Font(name="Arial", size=9, italic=True, color="555555"),
                alignment=CENTER)
    ws.row_dimensions[2].height = 18

    # Blank row
    ws.row_dimensions[3].height = 6

    # Headers
    headers = ["#", "Speaker Label", "Original Folder", "File Name", "Split"]
    col_widths = [6, 18, 22, 36, 10]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        _write_cell(ws, 4, col, h,
                    font=HEADER_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=_border())
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22

    # Data rows
    prev_spk = None
    for i, row in enumerate(rows, 1):
        excel_row = i + 4
        is_new_spk = (row["Speaker Label"] != prev_spk)
        prev_spk   = row["Speaker Label"]

        row_fill  = TRAIN_FILL if row["Split"] == "Train" else TEST_FILL
        spk_fill  = PatternFill("solid", fgColor="EEF2FF") \
                    if row["Split"] == "Train" else \
                    PatternFill("solid", fgColor="FFF8E1")

        # Stripe alternate speakers slightly
        base_fill = row_fill

        _write_cell(ws, excel_row, 1, i,
                    font=BODY_FONT, fill=base_fill,
                    alignment=CENTER, border=_border())
        _write_cell(ws, excel_row, 2, row["Speaker Label"],
                    font=BOLD_FONT if is_new_spk else BODY_FONT,
                    fill=base_fill, alignment=LEFT, border=_border())
        _write_cell(ws, excel_row, 3, row["Original Folder"],
                    font=BODY_FONT, fill=base_fill,
                    alignment=LEFT, border=_border())
        _write_cell(ws, excel_row, 4, row["File Name"],
                    font=BODY_FONT, fill=base_fill,
                    alignment=LEFT, border=_border())

        split_font = Font(name="Arial", size=10, bold=True,
                          color="1B5E20" if row["Split"]=="Train" else "0D47A1")
        _write_cell(ws, excel_row, 5, row["Split"],
                    font=split_font, fill=base_fill,
                    alignment=CENTER, border=_border())

        ws.row_dimensions[excel_row].height = 16

    # Freeze panes
    ws.freeze_panes = "A5"


# ── Sheet 2 – Per-Speaker Summary ─────────────────────────────────────────────

def build_summary_sheet(wb, rows, le):
    ws = wb.create_sheet("Per-Speaker Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:F1")
    _write_cell(ws, 1, 1, "Per-Speaker Train / Test Summary",
                font=TITLE_FONT, alignment=CENTER)
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 6

    headers = ["Speaker Label", "Original Folder",
               "Total Files", "Train Files", "Test Files", "Train %"]
    col_widths = [18, 22, 14, 14, 14, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        _write_cell(ws, 3, col, h,
                    font=HEADER_FONT, fill=HEADER_FILL,
                    alignment=CENTER, border=_border())
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 22

    # Aggregate per speaker
    from collections import defaultdict
    spk_data = defaultdict(lambda: {"folder": "", "train": 0, "test": 0})
    for r in rows:
        spk_data[r["Speaker Label"]]["folder"] = r["Original Folder"]
        if r["Split"] == "Train":
            spk_data[r["Speaker Label"]]["train"] += 1
        else:
            spk_data[r["Speaker Label"]]["test"] += 1

    for i, (spk, d) in enumerate(sorted(spk_data.items()), 1):
        excel_row = i + 3
        total = d["train"] + d["test"]
        fill  = PatternFill("solid", fgColor="F9FBE7" if i % 2 == 0 else "FFFFFF")

        _write_cell(ws, excel_row, 1, spk,
                    font=BOLD_FONT, fill=fill, alignment=LEFT, border=_border())
        _write_cell(ws, excel_row, 2, d["folder"],
                    font=BODY_FONT, fill=fill, alignment=LEFT, border=_border())
        _write_cell(ws, excel_row, 3, total,
                    font=BODY_FONT, fill=fill, alignment=CENTER, border=_border())
        _write_cell(ws, excel_row, 4, d["train"],
                    font=Font(name="Arial", size=10, color="1B5E20"),
                    fill=fill, alignment=CENTER, border=_border())
        _write_cell(ws, excel_row, 5, d["test"],
                    font=Font(name="Arial", size=10, color="0D47A1"),
                    fill=fill, alignment=CENTER, border=_border())
        # Train % formula
        pct_cell = ws.cell(row=excel_row, column=6)
        pct_cell.value = f"=D{excel_row}/C{excel_row}"
        pct_cell.font  = BODY_FONT
        pct_cell.fill  = fill
        pct_cell.alignment = CENTER
        pct_cell.border    = _border()
        pct_cell.number_format = "0.0%"

        ws.row_dimensions[excel_row].height = 16

    # Totals row
    total_row = len(spk_data) + 4
    fill_tot  = PatternFill("solid", fgColor="1F3864")
    font_tot  = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    _write_cell(ws, total_row, 1, "TOTAL", font=font_tot, fill=fill_tot,
                alignment=CENTER, border=_border())
    _write_cell(ws, total_row, 2, f"{len(le.classes_)} speakers",
                font=font_tot, fill=fill_tot, alignment=CENTER, border=_border())

    last_data = total_row - 1
    for col, formula in [
        (3, f"=SUM(C4:C{last_data})"),
        (4, f"=SUM(D4:D{last_data})"),
        (5, f"=SUM(E4:E{last_data})"),
        (6, f"=D{total_row}/C{total_row}"),
    ]:
        c = ws.cell(row=total_row, column=col, value=formula)
        c.font   = font_tot
        c.fill   = fill_tot
        c.alignment = CENTER
        c.border = _border()
        if col == 6:
            c.number_format = "0.0%"

    ws.row_dimensions[total_row].height = 20
    ws.freeze_panes = "A4"


# ── Sheet 3 – Train Only / Test Only ──────────────────────────────────────────

def build_split_sheets(wb, rows):
    for split_name, fill_color, font_color in [
        ("Train Files", "E8F5E9", "1B5E20"),
        ("Test Files",  "E3F2FD", "0D47A1"),
    ]:
        ws  = wb.create_sheet(split_name)
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D1")
        _write_cell(ws, 1, 1, split_name,
                    font=TITLE_FONT, alignment=CENTER)
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 6

        headers    = ["#", "Speaker Label", "Original Folder", "File Name"]
        col_widths = [6, 18, 22, 36]
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            _write_cell(ws, 3, col, h,
                        font=HEADER_FONT, fill=HEADER_FILL,
                        alignment=CENTER, border=_border())
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[3].height = 22

        split_key = split_name.split()[0]   # "Train" or "Test"
        filtered  = [r for r in rows if r["Split"] == split_key]

        fill = PatternFill("solid", fgColor=fill_color)
        for i, row in enumerate(filtered, 1):
            excel_row = i + 3
            _write_cell(ws, excel_row, 1, i,
                        font=BODY_FONT, fill=fill,
                        alignment=CENTER, border=_border())
            _write_cell(ws, excel_row, 2, row["Speaker Label"],
                        font=BOLD_FONT, fill=fill,
                        alignment=LEFT, border=_border())
            _write_cell(ws, excel_row, 3, row["Original Folder"],
                        font=BODY_FONT, fill=fill,
                        alignment=LEFT, border=_border())
            _write_cell(ws, excel_row, 4, row["File Name"],
                        font=BODY_FONT, fill=fill,
                        alignment=LEFT, border=_border())
            ws.row_dimensions[excel_row].height = 16

        ws.freeze_panes = "A4"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Collecting dataset records …")
    rows, le = get_splits()

    total  = len(rows)
    n_tr   = sum(1 for r in rows if r["Split"] == "Train")
    n_te   = total - n_tr

    print(f"  Total files : {total}")
    print(f"  Train       : {n_tr}  ({n_tr/total*100:.1f}%)")
    print(f"  Test        : {n_te}  ({n_te/total*100:.1f}%)")
    print(f"  Speakers    : {len(le.classes_)}")

    print("\nBuilding Excel workbook …")
    wb = Workbook()

    build_detail_sheet(wb, rows, le)
    build_summary_sheet(wb, rows, le)
    build_split_sheets(wb, rows)

    out_path = os.path.join(config.RESULTS_DIR, "dataset_splits.xlsx")
    wb.save(out_path)
    print(f"\n✅ Saved → {out_path}")
    print("   Sheets: Full Split Detail | Per-Speaker Summary | Train Files | Test Files")


if __name__ == "__main__":
    main()