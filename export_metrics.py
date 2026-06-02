"""
export_metrics.py  –  Documentation-grade metrics Excel
Sheets:
  1. Dashboard         – KPI cards + accuracy bar + macro F1 pie
  2. Speaker Confidence– heat-map table + clustered bar chart
  3. Model Report      – full P/R/F1/Support per speaker per model + horizontal bar
  4. PRF Graphs        – per-model line charts + combined F1 overlay
  5. Heatmap Grid      – colour-graded P / R / F1 grid (all models × all speakers)

Run: python export_metrics.py
"""

import os, sys, pickle
import numpy as np
import torch
from sklearn.metrics import precision_recall_fscore_support
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

import config
from data_loader  import load_ml_dataset, load_dl_dataset
from ml_models    import load_ml_models, load_label_encoder
from dl_models    import XVectorNet, ECAPATDNNNet, load_dl_model, get_device
from evaluate     import evaluate_ml, evaluate_dl


# ══════════════════════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

C_NAVY   = "1B2A4A";  C_BLUE   = "2563EB";  C_TEAL   = "0D9488"
C_GREEN  = "16A34A";  C_AMBER  = "D97706";  C_RED    = "DC2626"
C_PURPLE = "7C3AED";  C_ORANGE = "EA580C";  C_LGREY  = "F8FAFC"
C_MGREY  = "E2E8F0";  C_DGREY  = "64748B";  C_WHITE  = "FFFFFF"
C_BLACK  = "0F172A"

MODEL_COLORS = {
    "SVM":               "2563EB",
    "Random Forest":     "16A34A",
    "Gradient Boosting": "EA580C",
    "X-Vector":          "7C3AED",
    "ECAPA-TDNN":        "DC2626",
}
MODEL_LIGHT = {
    "SVM":               "DBEAFE",
    "Random Forest":     "DCFCE7",
    "Gradient Boosting": "FFEDD5",
    "X-Vector":          "EDE9FE",
    "ECAPA-TDNN":        "FEE2E2",
}


def hf(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def sd(color="CCCCCC", style="thin"):
    return Side(style=style, color=color)

def border(color="CCCCCC"):
    s = sd(color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border(color="888888"):
    s = sd(color, "medium")
    return Border(left=s, right=s, top=s, bottom=s)

def fnt(bold=False, sz=10, color=C_BLACK, italic=False):
    return Font(name="Arial", bold=bold, size=sz, color=color, italic=italic)

def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, r, c, val="", bg=C_WHITE, fg=C_BLACK,
         bold=False, sz=10, italic=False,
         h="center", v="center", fmt=None, brd=True):
    cc = ws.cell(row=r, column=c, value=val)
    cc.font      = fnt(bold=bold, sz=sz, color=fg, italic=italic)
    cc.fill      = hf(bg)
    cc.alignment = aln(h=h, v=v)
    if brd:
        cc.border = border()
    if fmt:
        cc.number_format = fmt
    return cc

def merge_title(ws, r, c1, c2, text, bg=C_NAVY, fg=C_WHITE, sz=13, h=30):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cc = ws.cell(row=r, column=c1, value=text)
    cc.font      = fnt(bold=True, sz=sz, color=fg)
    cc.fill      = hf(bg)
    cc.alignment = aln()
    ws.row_dimensions[r].height = h
    return cc

def col_w(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def row_h(ws, row_idx, height):
    ws.row_dimensions[row_idx].height = height

def heat_color(value, lo=0.0, hi=1.0):
    """Red → Amber → Green gradient for 0–1 values."""
    t = max(0.0, min(1.0, (value - lo) / max(hi - lo, 1e-9)))
    if t < 0.5:
        r = 220; g = int(87 + t * 2 * 166); b = 87
    else:
        r = int(220 - (t - 0.5) * 2 * 220); g = 163; b = 74
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    return f"{r:02X}{g:02X}{b:02X}"


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_and_evaluate():
    results = []

    # ML
    try:
        pipes = load_ml_models()
        le    = load_label_encoder()
        _, X_te, _, y_te, _ = load_ml_dataset()
        res = evaluate_ml(pipes, X_te, y_te, le)
        for mname, r in res.items():
            p, rec, f1, sup = precision_recall_fscore_support(
                r["y_true"], r["y_pred"], zero_division=0)
            probas = pipes[mname].predict_proba(X_te)
            conf   = {cls: float(probas[y_te==i, i].mean())*100
                      if (y_te==i).sum() > 0 else 0.0
                      for i, cls in enumerate(le.classes_)}
            results.append(dict(model=mname, type="ML",
                                accuracy=r["accuracy"], classes=le.classes_,
                                precision=p, recall=rec, f1=f1, support=sup,
                                conf_per_spk=conf))
    except Exception as e:
        print(f"  [WARN] ML: {e}")

    # DL
    try:
        with open(config.DL_LABEL_PATH, "rb") as f:
            le_dl = pickle.load(f)
        num_spk = len(le_dl.classes_)
        _, X_te_dl, _, y_te_dl, _ = load_dl_dataset()
        device = get_device()
        for MCls, path, mname in [
            (XVectorNet,   config.DL_XVEC_PATH,  "X-Vector"),
            (ECAPATDNNNet, config.DL_ECAPA_PATH,  "ECAPA-TDNN"),
        ]:
            mdl = load_dl_model(MCls, path, config.N_MELS, num_spk)
            rr  = evaluate_dl(mdl, X_te_dl, y_te_dl, le_dl, mname)[mname]
            p, rec, f1, sup = precision_recall_fscore_support(
                rr["y_true"], rr["y_pred"], zero_division=0)
            mdl.eval()
            with torch.no_grad():
                logits, _ = mdl(torch.tensor(X_te_dl, dtype=torch.float32).to(device))
            probs = torch.softmax(logits, 1).cpu().numpy()
            conf  = {cls: float(probs[y_te_dl==i, i].mean())*100
                     if (y_te_dl==i).sum() > 0 else 0.0
                     for i, cls in enumerate(le_dl.classes_)}
            results.append(dict(model=mname, type="DL",
                                accuracy=rr["accuracy"], classes=le_dl.classes_,
                                precision=p, recall=rec, f1=f1, support=sup,
                                conf_per_spk=conf))
    except Exception as e:
        print(f"  [WARN] DL: {e}")

    return results


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

def build_dashboard(wb, results):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Banner
    n_banner = 2 + len(results) * 3
    merge_title(ws, 1, 1, n_banner,
                "ASSAMESE SPEAKER IDENTIFICATION & VERIFICATION  —  Model Performance Dashboard",
                bg=C_NAVY, sz=14, h=38)
    merge_title(ws, 2, 1, n_banner,
                f"Models: {len(results)}   |   Speakers: {len(results[0]['classes'])}   |   "
                f"Language: Assamese   |   Train/Test Split: 80% / 20%   |   Random State: 42",
                bg=C_DGREY, fg=C_WHITE, sz=9, h=18)
    row_h(ws, 3, 10)

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    card_col = 1
    for r in results:
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        lclr = MODEL_LIGHT.get(r["model"], "EFF6FF")
        f1m  = float(np.mean(r["f1"]))
        pm   = float(np.mean(r["precision"]))
        rm   = float(np.mean(r["recall"]))

        # Model name
        ws.merge_cells(start_row=4, start_column=card_col,
                       end_row=4,   end_column=card_col+1)
        cc = ws.cell(row=4, column=card_col, value=r["model"])
        cc.font = fnt(bold=True, sz=10, color=C_WHITE)
        cc.fill = hf(mclr); cc.alignment = aln()
        cc.border = thick_border(mclr)

        # Type
        ws.merge_cells(start_row=5, start_column=card_col,
                       end_row=5,   end_column=card_col+1)
        cc2 = ws.cell(row=5, column=card_col, value=f"Type: {r['type']}")
        cc2.font = fnt(sz=8, color=mclr, italic=True)
        cc2.fill = hf(lclr); cc2.alignment = aln()

        # Accuracy
        cell(ws, 6, card_col,   "Accuracy",    bg=lclr, fg=C_DGREY, sz=8)
        c4 = ws.cell(row=6, column=card_col+1, value=r["accuracy"])
        c4.font = fnt(bold=True, sz=12, color=mclr)
        c4.fill = hf(lclr); c4.alignment = aln(); c4.number_format = "0.0%"

        # Macro F1
        cell(ws, 7, card_col,   "Macro F1",    bg=C_LGREY, fg=C_DGREY, sz=8)
        c5 = ws.cell(row=7, column=card_col+1, value=f1m)
        c5.font = fnt(bold=True, sz=12, color=mclr)
        c5.fill = hf(C_LGREY); c5.alignment = aln(); c5.number_format = "0.0%"

        # Macro Precision
        cell(ws, 8, card_col,   "Macro P",     bg=lclr, fg=C_DGREY, sz=8)
        c6 = ws.cell(row=8, column=card_col+1, value=pm)
        c6.font = fnt(sz=10, color=mclr)
        c6.fill = hf(lclr); c6.alignment = aln(); c6.number_format = "0.0%"

        # Macro Recall
        cell(ws, 9, card_col,   "Macro R",     bg=C_LGREY, fg=C_DGREY, sz=8)
        c7 = ws.cell(row=9, column=card_col+1, value=rm)
        c7.font = fnt(sz=10, color=mclr)
        c7.fill = hf(C_LGREY); c7.alignment = aln(); c7.number_format = "0.0%"

        for rr in [4, 5, 6, 7, 8, 9]:
            row_h(ws, rr, 20)

        card_col += 3

    # Column widths for cards
    for i in range(1, card_col + 2):
        col_w(ws, i, 11)

    row_h(ws, 10, 10)

    # ── Summary table ─────────────────────────────────────────────────────────
    merge_title(ws, 11, 1, 10,
                "Model Performance Summary Table",
                bg=C_BLUE, sz=11, h=24)

    hdrs = ["Model", "Type", "Accuracy", "Macro P", "Macro R",
            "Macro F1", "Weighted F1", "Avg Confidence", "Support"]
    widths = [22, 8, 12, 12, 12, 12, 14, 16, 11]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        cell(ws, 12, ci, h, bg=C_NAVY, fg=C_WHITE, bold=True, sz=9)
        col_w(ws, ci, w)
    row_h(ws, 12, 20)

    data_start = 13
    for i, r in enumerate(results):
        er   = data_start + i
        bg   = C_LGREY if i % 2 == 0 else C_WHITE
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        pm   = float(np.mean(r["precision"]))
        rm   = float(np.mean(r["recall"]))
        fm   = float(np.mean(r["f1"]))
        tot  = int(r["support"].sum())
        fw   = float(np.sum(r["f1"] * r["support"]) / tot) if tot else 0
        ac   = np.mean(list(r["conf_per_spk"].values()))
        acc_c = C_GREEN if r["accuracy"] >= 0.85 else (
                C_AMBER if r["accuracy"] >= 0.70 else C_RED)

        cell(ws, er, 1, r["model"],       bg=bg, fg=mclr,  bold=True, h="left")
        cell(ws, er, 2, r["type"],        bg=bg)
        cell(ws, er, 3, r["accuracy"],    bg=bg, fg=acc_c, bold=True, fmt="0.0%")
        cell(ws, er, 4, pm,               bg=bg, fmt="0.0%")
        cell(ws, er, 5, rm,               bg=bg, fmt="0.0%")
        cell(ws, er, 6, fm,               bg=bg, fmt="0.0%", bold=True)
        cell(ws, er, 7, fw,               bg=bg, fmt="0.0%")
        cell(ws, er, 8, ac/100,           bg=bg, fmt="0.0%")
        cell(ws, er, 9, tot,              bg=bg)
        row_h(ws, er, 17)

    last_data = data_start + len(results) - 1
    best_idx  = max(range(len(results)), key=lambda i: results[i]["accuracy"])
    best_er   = last_data + 1
    merge_title(ws, best_er, 1, 9,
                f"🏆  Best Model: {results[best_idx]['model']}  "
                f"({results[best_idx]['accuracy']*100:.1f}% accuracy)  "
                f"|  Macro F1: {np.mean(results[best_idx]['f1'])*100:.1f}%",
                bg=MODEL_COLORS.get(results[best_idx]["model"], C_GREEN),
                sz=10, h=22)

    ws.freeze_panes = "A13"

    # ── Accuracy Bar Chart ────────────────────────────────────────────────────
    chart_r = best_er + 3
    chart = BarChart()
    chart.type   = "col"; chart.title  = "Model Accuracy Comparison"
    chart.y_axis.title = "Accuracy"; chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0;    chart.y_axis.scaling.max = 1
    chart.x_axis.title = "Model";    chart.legend = None
    chart.width = 18; chart.height = 13; chart.style = 2

    dr = Reference(ws, min_col=3, min_row=12, max_row=last_data)
    cr = Reference(ws, min_col=1, min_row=data_start, max_row=last_data)
    chart.add_data(dr, titles_from_data=True)
    chart.set_categories(cr)
    for j, r in enumerate(results):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = MODEL_COLORS.get(r["model"], C_BLUE)
        chart.series[0].dPt.append(pt)
    ws.add_chart(chart, f"A{chart_r}")

    # ── F1 Pie Chart ──────────────────────────────────────────────────────────
    pie_r = chart_r
    for j, r in enumerate(results):
        ws.cell(row=pie_r+j, column=11, value=r["model"])
        ws.cell(row=pie_r+j, column=12, value=float(np.mean(r["f1"])))

    pie = PieChart()
    pie.title = "Macro F1 Distribution by Model"
    pie.width = 16; pie.height = 13; pie.style = 10

    pd_ref = Reference(ws, min_col=12, min_row=pie_r, max_row=pie_r+len(results)-1)
    pc_ref = Reference(ws, min_col=11, min_row=pie_r, max_row=pie_r+len(results)-1)
    pie.add_data(pd_ref); pie.set_categories(pc_ref)
    for j, r in enumerate(results):
        pt = DataPoint(idx=j)
        pt.graphicalProperties.solidFill = MODEL_COLORS.get(r["model"], C_BLUE)
        pie.series[0].dPt.append(pt)
    ws.add_chart(pie, f"K{chart_r}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – SPEAKER CONFIDENCE HEATMAP
# ══════════════════════════════════════════════════════════════════════════════

def build_confidence(wb, results):
    ws = wb.create_sheet("Speaker Confidence")
    ws.sheet_view.showGridLines = False

    classes  = results[0]["classes"]
    n_cols   = len(results) + 2

    merge_title(ws, 1, 1, n_cols,
                "Speaker-wise Confidence (%)  —  Heat-Map View",
                bg=C_TEAL, sz=13, h=32)
    merge_title(ws, 2, 1, n_cols,
                "Green = High confidence (≥80%)     Amber = Medium (60–79%)     Red = Low (<60%)",
                bg="134E4A", fg="99F6E4", sz=8, h=16)
    row_h(ws, 3, 8)

    col_w(ws, 1, 6); col_w(ws, 2, 18)
    cell(ws, 4, 1, "#",       bg=C_TEAL, fg=C_WHITE, bold=True, sz=9)
    cell(ws, 4, 2, "Speaker", bg=C_TEAL, fg=C_WHITE, bold=True, sz=9)
    for j, r in enumerate(results, 3):
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        cell(ws, 4, j, r["model"], bg=mclr, fg=C_WHITE, bold=True, sz=9)
        col_w(ws, j, 18)
    row_h(ws, 4, 22)

    all_confs = [list(r["conf_per_spk"].values()) for r in results]
    glo = min(min(c) for c in all_confs)
    ghi = max(max(c) for c in all_confs)

    for i, cls in enumerate(classes):
        er  = i + 5
        rbg = C_LGREY if i % 2 == 0 else C_WHITE
        cell(ws, er, 1, i+1, bg=rbg, sz=9)
        cell(ws, er, 2, cls, bg=rbg, bold=True, sz=9, h="left")
        for j, r in enumerate(results, 3):
            v    = r["conf_per_spk"].get(cls, 0.0)
            heat = heat_color(v, glo, ghi)
            txt  = C_WHITE if v < 70 else C_BLACK
            cell(ws, er, j, v/100, bg=heat, fg=txt,
                 bold=(v >= 80), fmt="0.0%", sz=10)
        row_h(ws, er, 17)

    avg_er = len(classes) + 5
    cell(ws, avg_er, 1, "",        bg=C_NAVY)
    cell(ws, avg_er, 2, "AVERAGE", bg=C_NAVY, fg=C_WHITE, bold=True)
    for j, r in enumerate(results, 3):
        avg  = np.mean(list(r["conf_per_spk"].values()))
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        cell(ws, avg_er, j, avg/100,
             bg=mclr, fg=C_WHITE, bold=True, fmt="0.0%", sz=11)
    row_h(ws, avg_er, 22)

    ws.freeze_panes = "C5"

    # Clustered bar chart
    data_end = len(classes) + 4
    chart = BarChart()
    chart.type = "col"; chart.grouping = "clustered"
    chart.title = "Speaker-wise Confidence by Model"
    chart.y_axis.title = "Confidence"; chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0;     chart.y_axis.scaling.max = 1
    chart.x_axis.title = "Speaker"
    chart.width = 36; chart.height = 18

    for j, r in enumerate(results, 3):
        dr = Reference(ws, min_col=j, max_col=j, min_row=4, max_row=data_end)
        chart.add_data(dr, titles_from_data=True)
        clr = MODEL_COLORS.get(r["model"], C_BLUE)
        chart.series[j-3].graphicalProperties.solidFill = clr
        chart.series[j-3].graphicalProperties.line.solidFill = clr

    chart.set_categories(Reference(ws, min_col=2, min_row=5, max_row=data_end))
    ws.add_chart(chart, f"A{avg_er + 3}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – MODEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

def build_model_report(wb, results):
    ws = wb.create_sheet("Model Report")
    ws.sheet_view.showGridLines = False

    classes  = results[0]["classes"]
    n_models = len(results)
    total_cols = 1 + n_models * 4 + 1

    merge_title(ws, 1, 1, total_cols,
                "Per-Speaker Classification Report  —  Precision  |  Recall  |  F1-Score  |  Support",
                bg=C_NAVY, sz=13, h=32)
    merge_title(ws, 2, 1, total_cols,
                "Colour coding:  Green ≥85%     Amber ≥65%     Red <65%     Bold = strong performer",
                bg=C_DGREY, fg=C_MGREY, sz=8, h=16)
    row_h(ws, 3, 8)

    col_w(ws, 1, 6); col_w(ws, 2, 18)
    cell(ws, 4, 1, "#",       bg="37474F", fg=C_WHITE, bold=True)
    cell(ws, 4, 2, "Speaker", bg="37474F", fg=C_WHITE, bold=True)

    # Model group headers
    for j, r in enumerate(results):
        base = 3 + j * 4
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        ws.merge_cells(start_row=4, start_column=base,
                       end_row=4,   end_column=base+3)
        cc = ws.cell(row=4, column=base, value=r["model"])
        cc.font = fnt(bold=True, sz=10, color=C_WHITE)
        cc.fill = hf(mclr); cc.alignment = aln()
        for k in range(4):
            col_w(ws, base+k, 13)
    row_h(ws, 4, 22)

    # Sub-headers
    cell(ws, 5, 1, "", bg=C_MGREY); cell(ws, 5, 2, "", bg=C_MGREY)
    for j, r in enumerate(results):
        base = 3 + j * 4
        lclr = MODEL_LIGHT.get(r["model"], "EFF6FF")
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        for k, lbl in enumerate(["Precision", "Recall", "F1", "Support"]):
            cell(ws, 5, base+k, lbl, bg=lclr, fg=mclr, bold=True, sz=8)
    row_h(ws, 5, 18)

    # Data rows
    for i, cls in enumerate(classes):
        er = i + 6
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        cell(ws, er, 1, i+1, bg=bg, sz=9)
        cell(ws, er, 2, cls, bg=bg, bold=True, h="left", sz=9)
        for j, r in enumerate(results):
            base = 3 + j * 4
            pv = float(r["precision"][i]); rv = float(r["recall"][i])
            fv = float(r["f1"][i]);        sv = int(r["support"][i])
            pc = C_GREEN if pv>=0.85 else (C_AMBER if pv>=0.65 else C_RED)
            rc = C_GREEN if rv>=0.85 else (C_AMBER if rv>=0.65 else C_RED)
            fc = C_GREEN if fv>=0.85 else (C_AMBER if fv>=0.65 else C_RED)
            cell(ws, er, base,   pv, bg=bg, fg=pc, fmt="0.0%", bold=(pv>=0.85))
            cell(ws, er, base+1, rv, bg=bg, fg=rc, fmt="0.0%", bold=(rv>=0.85))
            cell(ws, er, base+2, fv, bg=bg, fg=fc, fmt="0.0%", bold=(fv>=0.85))
            cell(ws, er, base+3, sv, bg=bg, sz=9)
        row_h(ws, er, 16)

    # Macro avg
    avg_er = len(classes) + 6
    cell(ws, avg_er, 1, "",           bg=C_NAVY)
    cell(ws, avg_er, 2, "MACRO AVG",  bg=C_NAVY, fg=C_WHITE, bold=True)
    for j, r in enumerate(results):
        base = 3 + j * 4
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        for k, arr in enumerate([r["precision"], r["recall"], r["f1"]]):
            cc = ws.cell(row=avg_er, column=base+k, value=float(np.mean(arr)))
            cc.font = fnt(bold=True, sz=10, color=C_WHITE)
            cc.fill = hf(mclr); cc.alignment = aln()
            cc.border = border(); cc.number_format = "0.0%"
        cc = ws.cell(row=avg_er, column=base+3, value=int(r["support"].sum()))
        cc.font = fnt(bold=True, color=C_WHITE)
        cc.fill = hf(mclr); cc.alignment = aln(); cc.border = border()
    row_h(ws, avg_er, 20)
    ws.freeze_panes = "C6"

    # Horizontal bar chart (Macro P / R / F1)
    stg = total_cols + 2
    ws.cell(row=3, column=stg,   value="Model")
    ws.cell(row=3, column=stg+1, value="Precision")
    ws.cell(row=3, column=stg+2, value="Recall")
    ws.cell(row=3, column=stg+3, value="F1")
    for j, r in enumerate(results, 1):
        ws.cell(row=3+j, column=stg,   value=r["model"])
        ws.cell(row=3+j, column=stg+1, value=float(np.mean(r["precision"])))
        ws.cell(row=3+j, column=stg+2, value=float(np.mean(r["recall"])))
        ws.cell(row=3+j, column=stg+3, value=float(np.mean(r["f1"])))

    chart = BarChart()
    chart.type = "bar"; chart.grouping = "clustered"
    chart.title = "Macro Precision / Recall / F1 by Model"
    chart.x_axis.title = "Score"; chart.x_axis.numFmt = "0%"
    chart.x_axis.scaling.min = 0; chart.x_axis.scaling.max = 1
    chart.y_axis.title = "Model"
    chart.width = 22; chart.height = 14
    for k, (col_off, clr) in enumerate([(1, "2563EB"), (2, "EA580C"), (3, "16A34A")]):
        dr = Reference(ws, min_col=stg+col_off, max_col=stg+col_off,
                       min_row=3, max_row=3+len(results))
        chart.add_data(dr, titles_from_data=True)
        chart.series[k].graphicalProperties.solidFill = clr
    chart.set_categories(
        Reference(ws, min_col=stg, min_row=4, max_row=3+len(results)))
    ws.add_chart(chart, f"A{avg_er + 3}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – PRF LINE GRAPHS
# ══════════════════════════════════════════════════════════════════════════════

def build_prf_graphs(wb, results):
    ws = wb.create_sheet("PRF Graphs")
    ws.sheet_view.showGridLines = False

    classes = results[0]["classes"]
    n_cls   = len(classes)
    n_cols  = 1 + len(results) * 3

    merge_title(ws, 1, 1, n_cols,
                "Precision / Recall / F1  per Speaker  —  All Models",
                bg=C_NAVY, sz=13, h=32)
    merge_title(ws, 2, 1, n_cols,
                "Blue line = Precision     Orange line = Recall     Green line = F1-Score",
                bg=C_DGREY, fg=C_WHITE, sz=8, h=16)
    row_h(ws, 3, 8)

    col_w(ws, 1, 18)
    cell(ws, 4, 1, "Speaker", bg=C_NAVY, fg=C_WHITE, bold=True)

    for j, r in enumerate(results):
        base = 2 + j * 3
        mclr = MODEL_COLORS.get(r["model"], C_BLUE)
        lclr = MODEL_LIGHT.get(r["model"], "EFF6FF")
        ws.merge_cells(start_row=4, start_column=base, end_row=4, end_column=base+2)
        cc = ws.cell(row=4, column=base, value=r["model"])
        cc.font = fnt(bold=True, sz=10, color=C_WHITE)
        cc.fill = hf(mclr); cc.alignment = aln()
        for k, lbl in enumerate(["Precision", "Recall", "F1"]):
            cell(ws, 5, base+k, lbl, bg=lclr, fg=mclr, bold=True, sz=8)
            col_w(ws, base+k, 14)
    row_h(ws, 4, 22); row_h(ws, 5, 18)

    for i, cls in enumerate(classes):
        er = i + 6
        bg = C_LGREY if i % 2 == 0 else C_WHITE
        cell(ws, er, 1, cls, bg=bg, bold=True, h="left", sz=9)
        for j, r in enumerate(results):
            base = 2 + j * 3
            pv = float(r["precision"][i]); rv = float(r["recall"][i])
            fv = float(r["f1"][i])
            cell(ws, er, base,   pv, bg=bg, fmt="0.0%",
                 fg=C_GREEN if pv>=0.85 else (C_AMBER if pv>=0.65 else C_RED))
            cell(ws, er, base+1, rv, bg=bg, fmt="0.0%",
                 fg=C_GREEN if rv>=0.85 else (C_AMBER if rv>=0.65 else C_RED))
            cell(ws, er, base+2, fv, bg=bg, fmt="0.0%",
                 fg=C_GREEN if fv>=0.85 else (C_AMBER if fv>=0.65 else C_RED),
                 bold=(fv>=0.85))
        row_h(ws, er, 15)

    data_end = n_cls + 5
    ws.freeze_panes = "B6"

    line_colors = ["2563EB", "EA580C", "16A34A"]

    # Per-model line charts
    chart_row = data_end + 3
    for j, r in enumerate(results):
        base  = 2 + j * 3
        chart = LineChart()
        chart.title = f"{r['model']}  —  Precision / Recall / F1 per Speaker"
        chart.y_axis.title = "Score"; chart.y_axis.numFmt = "0%"
        chart.y_axis.scaling.min = 0; chart.y_axis.scaling.max = 1
        chart.x_axis.title = "Speaker"
        chart.width = 26; chart.height = 14; chart.style = 10

        for k in range(3):
            dr = Reference(ws, min_col=base+k, max_col=base+k,
                           min_row=5, max_row=data_end)
            chart.add_data(dr, titles_from_data=True)
            s = chart.series[k]
            s.graphicalProperties.line.solidFill = line_colors[k]
            s.graphicalProperties.line.width = 18000
            s.marker.symbol = "circle"; s.marker.size = 5
            s.marker.graphicalProperties.solidFill = line_colors[k]
            s.marker.graphicalProperties.line.solidFill = line_colors[k]

        chart.set_categories(Reference(ws, min_col=1, min_row=6, max_row=data_end))
        col_off = "A" if j % 2 == 0 else "P"
        chart_r = chart_row + (j // 2) * 26
        ws.add_chart(chart, f"{col_off}{chart_r}")

    # Combined F1 overlay
    combined_r = chart_row + ((len(results) + 1) // 2) * 26 + 2
    merge_title(ws, combined_r - 2, 1, n_cols,
                "Combined F1-Score Comparison — All Models Overlaid",
                bg=C_TEAL, sz=11, h=22)

    chart_all = LineChart()
    chart_all.title = "F1-Score: All Models vs All Speakers"
    chart_all.y_axis.title = "F1 Score"; chart_all.y_axis.numFmt = "0%"
    chart_all.y_axis.scaling.min = 0; chart_all.y_axis.scaling.max = 1
    chart_all.x_axis.title = "Speaker"
    chart_all.width = 36; chart_all.height = 18; chart_all.style = 10

    for j, r in enumerate(results):
        f1_col = 2 + j * 3 + 2
        dr = Reference(ws, min_col=f1_col, max_col=f1_col,
                       min_row=5, max_row=data_end)
        chart_all.add_data(dr, titles_from_data=True)
        clr = MODEL_COLORS.get(r["model"], C_BLUE)
        s   = chart_all.series[j]
        s.graphicalProperties.line.solidFill = clr
        s.graphicalProperties.line.width = 22000
        s.marker.symbol = "circle"; s.marker.size = 6
        s.marker.graphicalProperties.solidFill = clr
        s.marker.graphicalProperties.line.solidFill = clr

    chart_all.set_categories(Reference(ws, min_col=1, min_row=6, max_row=data_end))
    ws.add_chart(chart_all, f"A{combined_r}")


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 – HEATMAP GRID
# ══════════════════════════════════════════════════════════════════════════════

def build_heatmap_grid(wb, results):
    ws = wb.create_sheet("Heatmap Grid")
    ws.sheet_view.showGridLines = False

    classes  = results[0]["classes"]
    n_models = len(results)
    n_cols   = n_models + 2

    merge_title(ws, 1, 1, n_cols,
                "Heatmap Grid  —  Precision  |  Recall  |  F1-Score   (Red = Low  →  Green = High)",
                bg=C_NAVY, sz=13, h=32)
    row_h(ws, 2, 8)

    metrics = [
        ("Precision", [r["precision"] for r in results], C_BLUE),
        ("Recall",    [r["recall"]    for r in results], C_PURPLE),
        ("F1-Score",  [r["f1"]        for r in results], C_TEAL),
    ]

    col_w(ws, 1, 6); col_w(ws, 2, 18)
    for j in range(3, n_models + 3):
        col_w(ws, j, 16)

    start_row = 3
    for metric_name, arrays, hdr_color in metrics:
        merge_title(ws, start_row, 1, n_cols,
                    f"◆   {metric_name}",
                    bg=hdr_color, sz=11, h=24)

        cell(ws, start_row+1, 1, "#",       bg="37474F", fg=C_WHITE, bold=True, sz=8)
        cell(ws, start_row+1, 2, "Speaker", bg="37474F", fg=C_WHITE, bold=True, sz=8)
        for j, r in enumerate(results, 3):
            mclr = MODEL_COLORS.get(r["model"], C_BLUE)
            cell(ws, start_row+1, j, r["model"], bg=mclr, fg=C_WHITE, bold=True, sz=8)
        row_h(ws, start_row+1, 20)

        all_vals = np.concatenate(arrays)
        lo, hi   = float(all_vals.min()), float(all_vals.max())

        for i, cls in enumerate(classes):
            er     = start_row + 2 + i
            row_bg = C_LGREY if i % 2 == 0 else C_WHITE
            cell(ws, er, 1, i+1, bg=row_bg, sz=8)
            cell(ws, er, 2, cls, bg=row_bg, bold=True, sz=8, h="left")
            vals_this_row = [float(arr[i]) for arr in arrays]
            best_val      = max(vals_this_row)
            for j, arr in enumerate(arrays, 3):
                v    = float(arr[i])
                heat = heat_color(v, lo, hi)
                txt  = C_WHITE if v < (lo + (hi - lo) * 0.55) else C_BLACK
                cell(ws, er, j, v, bg=heat, fg=txt, fmt="0.0%",
                     bold=(v == best_val), sz=9)
            row_h(ws, er, 16)

        # Avg row
        avg_er = start_row + 2 + len(classes)
        cell(ws, avg_er, 1, "",    bg=C_NAVY)
        cell(ws, avg_er, 2, "AVG", bg=C_NAVY, fg=C_WHITE, bold=True, sz=9)
        for j, (r, arr) in enumerate(zip(results, arrays), 3):
            mclr = MODEL_COLORS.get(r["model"], C_BLUE)
            cell(ws, avg_er, j, float(np.mean(arr)),
                 bg=mclr, fg=C_WHITE, bold=True, fmt="0.0%", sz=10)
        row_h(ws, avg_er, 20)

        start_row = avg_er + 3

    ws.freeze_panes = "C4"


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Loading & evaluating models …")
    results = load_and_evaluate()

    if not results:
        print("No models found. Run:  python train.py --mode all")
        return

    print(f"  Models   : {[r['model'] for r in results]}")
    print(f"  Speakers : {len(results[0]['classes'])}")

    wb = Workbook()
    print("  Sheet 1 : Dashboard …")
    build_dashboard(wb, results)
    print("  Sheet 2 : Speaker Confidence …")
    build_confidence(wb, results)
    print("  Sheet 3 : Model Report …")
    build_model_report(wb, results)
    print("  Sheet 4 : PRF Graphs …")
    build_prf_graphs(wb, results)
    print("  Sheet 5 : Heatmap Grid …")
    build_heatmap_grid(wb, results)

    out = os.path.join(config.RESULTS_DIR, "model_metrics.xlsx")
    wb.save(out)
    print(f"\n✅  Saved → {out}")
    print("    Sheets: Dashboard | Speaker Confidence | Model Report | PRF Graphs | Heatmap Grid")


if __name__ == "__main__":
    main()